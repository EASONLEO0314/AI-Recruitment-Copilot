"""Build the local job knowledge base from the hiring spreadsheet."""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from backend.app.knowledge_base import build_knowledge_base, clean_text  # noqa: E402


DEFAULT_SOURCE = "/Users/eason/Documents/招聘信息表.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "backend" / "app" / "data" / "job_knowledge_base.json"


def load_rows(source: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required to build the knowledge base. "
            "Install backend dev requirements first."
        ) from exc

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=False))
    if not rows:
        return []

    headers = [clean_text(cell.value) for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row_index, cells in enumerate(rows[1:], start=2):
        record = {
            header: cell.value
            for header, cell in zip(headers, cells)
            if header
        }
        if not any(clean_text(value) for value in record.values()):
            continue
        record["_source_row"] = row_index
        records.append(record)
    return records


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate backend/app/data/job_knowledge_base.json from an xlsx file."
    )
    parser.add_argument(
        "source",
        nargs="?",
        default=os.environ.get("JOB_KB_SOURCE_XLSX", DEFAULT_SOURCE),
        help="Path to 招聘信息表.xlsx",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Output JSON path",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source = Path(args.source).expanduser()
    output = Path(args.output).expanduser()
    if not source.exists():
        raise SystemExit(f"source file not found: {source}")

    records = load_rows(source)
    kb = build_knowledge_base(
        records,
        source_name=source.name,
        generated_at=datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(kb, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        "generated "
        f"{output} from {source.name}: "
        f"{len(kb['jobs'])} jobs, {len(kb['concepts'])} concepts, "
        f"{len(kb['documents'])} documents"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
